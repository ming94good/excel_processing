import requests
import pandas as pd
import json
import sqlite3
from openpyxl import Workbook
import pymongo

sqlDB='new_info.db'
testCSV='test_log.csv'

# Mongo DB operation
def insert(user_id, recommend):
    if(count(user_id) > 0):
        for r in recommend:
            update(user_id, r)
    else:
        d={'user_id': user_id, 'recommend': recommend}
        return col.insert_one(d)
def count(user_id):
    return col.count_documents({"user_id": user_id})
def update(user_id, recommend):
    query = { "user_id": user_id}
    newvalues = { "$push": { "recommend": recommend } }
    return col.update_one(query, newvalues)
def search(user_id):
    query = { "user_id": user_id}
    return col.find(query)
def recommend(user_id):
    for x in search(user_id):
        return x['recommend']
def delete(user_id):
    query = { "user_id": user_id}
    return col.delete_many(query)
def clearData(col):
    col.remove()

# Fetch info from SQL DB(Train result)
def vtv(video_id):
    conn = sqlite3.connect(sqlDB)
    sql_cmd = f"SELECT recommended_videos FROM recommended_videos where video_id={video_id}"
    cur = conn.cursor()
    cur.execute(sql_cmd)
    videos = cur.fetchone()
    videos = [int(video) for video in videos[0].split(",")][:50]
    conn.close()
    return videos

# Read test data
df = pd.read_csv(testCSV)

# MongoDB setup
myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["mydatabase"]
col = mydb["recommend"]
clearData(col)

# Count True
c=0
# Count False
d=0

# Excel setup
wb = Workbook()
sheet = wb.create_sheet(title="video_recommend")
sheet.cell(row = 1, column = 1).value = 'user_id'
sheet.cell(row = 1, column = 2).value = 'video_id'
sheet.cell(row = 1, column = 3).value = 'video is in recommend list'
sheet.cell(row = 1, column = 4).value = 'True'
sheet.cell(row = 1, column = 5).value = 'False'
sheet.cell(row = 1, column = 6).value = 'Accuracy'

for index, row in df.iterrows():
    try:
        recommend_num = 50
        rmd = vtv(str(row['video_id']))
        id = int(row['user_id'])

        recommend_list = recommend(id) #Get recommend videos for the user
        rst = '#'
        if recommend_list:
            if row['video_id'] in recommend_list:
                c+=1
                rst = 'True'
            else:
                d+=1
                rst = 'False'
        insert(id,rmd) #Update recommend videos for the user
        
        sheet.cell(index+2, 1).value = str(row['user_id'])
        sheet.cell(index+2, 2).value = str(row['video_id'])
        sheet.cell(index+2, 3).value = str(rst)
    except Exception as e:
        print(e)
        continue
    
sheet.cell(2, 4).value = str(c)
sheet.cell(2, 5).value = str(d)
sheet.cell(2, 6).value = str(c/ (c+d))
wb.save('result.xlsx')
print(c)
print(d)
print(index)
print(c/ (c+d)) #Final accuracy
